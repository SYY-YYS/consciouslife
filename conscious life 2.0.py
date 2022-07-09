#Concious life (version 2.0) (can input and edit to-do items)

from time import *
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

def input_topic(): #2 uses /input items into a list
    while True:
        try:
            hmi = int(input('how many item to add:'))
            do = []
            for i in range(hmi):
                item_name = input('item name:')
                do.append(item_name)
            return do
        except:
            print('please input a number')


def Topic(dol):
    while True: #while loop within def:>
        topic = input('wt r u going to do\n"add" to add item:')
        if topic == 'add': #add item with checking repeated item
            dol1 = input_topic()
            print (dol1)
            re = 0
            for col in range(1,ws.max_column): #check if row 1 has the item
                chara = get_column_letter(col+1) #from column2 'B' to
                if ws[chara + '1'].value in dol1: #check if row has new item
                    print('repeated item')
                    re = 'repeat'
                    break
                else:
                    pass
            if re != 'repeat':
                for i in range(len(dol1)): #if upper no break,here will run
                    ws[get_column_letter(ws.max_column + 1) + '1'] = dol1[i]


            break
        elif topic not in dol:
            print('type again {} or add item'.format(dol))
            continue
        else:
            return topic
    return 0

def po():
    while True:
        proo = input('timer/ stop:')  # process operation
        if proo == 'stop':
            op_time = time()
            tt = op_time-st_time
            print('accumulated time = {}min'.format( int((tt) / 60)))
            return tt/60**2

        if proo == 'timer':
            op_time = time()
            tt = op_time-st_time
            print('accumulated time = {}min'.format(int((tt) / 60)))
            continue
        else:
            print('type timer/stop')
            continue

def open_excel(excel_name):
    try:
        wb = load_workbook(excel_name)
        ws = wb.active
    except:
        wb = Workbook()
        ws = wb.active
    return wb, ws

def according_topic1(topic,row,tt):
    for i in range(1, ws.max_column): #locate column for selected item
        chara = get_column_letter(i+1)
        if ws[chara + '1'].value == topic: #locate column for selected item
            if ws[chara + str(row)].value != None:  # check if any record before
                ws[chara + str(row)] = ws[chara + str(row)].value + '+' + tt
            else:
                ws[chara + str(row)] = '=' + tt
        else:
            pass

while True:
    wb, ws = open_excel('conscious life1.xlsx')
    if ws['a1'].value == None:
        dol = input_topic()
        dol = ['date DD MM YYYY'] + dol
        for i in range(1,len(dol)+1):
            chara = get_column_letter(i)
            ws[chara + '1'] = dol[i-1] #checked
    else:
        dol =[]
        for col in range(1,ws.max_column):
            chara = get_column_letter(col+1) #col 1 is date
            dol.append(ws[chara + '1'].value)

    topic = Topic(dol) #dol stopped so no need to save dol2

    st_time = time()
    if topic != 0:
        tt = str(po())
    else:
        tt = 0
    #save to excel
    t = gmtime() #date would be the day you open the program
    td = str(t[2])+' '+str(t[1])+' '+str(t[0])

    for row in range(1, ws.max_row+1):
        if ws['a' + str(row)].value == td: #step 1:check date (check if there is, if not then add)
            according_topic1(topic,row,tt) #select column (default: study to column b)
        elif ws['a' + str(ws.max_row)].value != td:
            ws.append([td])
            according_topic1(topic,ws.max_row,tt)
        else:
            pass
    wb.save('conscious life1.xlsx')
